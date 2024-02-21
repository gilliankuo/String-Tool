import sys
import os
import openpyxl
import re

app_key_identifier = "App key"
language_identifiers = ["zh_TW", "JA", "EN", "TH", "zh_HK", "zh_CN"]

variable_identifier = "App variable"
valid_variable = {"%@", "%d"}


def read_sheets_workbook(file_path):
    wb = openpyxl.load_workbook(file_path)
    return wb


def generate_ios_strings(sheet, file, key_column, target_column):
    max_col = sheet.max_column
    for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
        if row[key_column] and row[target_column]:
            variable_placeholder = row[find_target_column(sheet, variable_identifier)]
            value = parse_string_value(row[target_column], variable_placeholder)

            ios_string = f'"{row[key_column]}" = "{value}";'
            file.write(ios_string + '\n')


def parse_string_value(input_string, variable_placeholder):
    parsed = input_string.replace("'", "\\'")
    parsed = parsed.replace("\n", "\\n")
    parsed = parsed.replace("<br/>", "\\n")
    parsed = parsed.strip()

    if variable_placeholder in valid_variable:
        parsed = re.sub(r'\{[^\}]*\}', variable_placeholder, parsed)

    return parsed



def remove_empty_source(input_string):
    pattern = r'"" = ".+?";'
    result = re.sub(pattern, '', input_string)
    return result


def remove_empty_row(input_string):
    result = re.sub(r'\n\s*\n', '\n', input_string)
    return result


def find_target_column(sheet, target):
    header_row = sheet[1]

    for cell in header_row:
        if cell.value and cell.value.lower().startswith(target.lower()):
            return cell.column - 1

    return -1


def process_sheet(sheet, file, key_column, target_column):
    # 產生 iOS 字串
    generate_ios_strings(sheet, file, key_column, target_column)


def main():
    # 檢查命令列參數
    if len(sys.argv) != 2:
        print("Please provide input file.")
        sys.exit(1)

    # 接受命令列輸入的 sheets_file_path
    sheets_file_path = sys.argv[1]

    # 獲取腳本目錄
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 創建一個資料夾用來存放輸出檔案
    output_folder_path = os.path.join(script_dir, 'output')
    os.makedirs(output_folder_path, exist_ok=True)

    # 讀取 Google Sheets
    workbook = read_sheets_workbook(sheets_file_path)

    # 開啟輸出檔案
    for language in language_identifiers:
        # 指定輸出檔案
        output_file_path = os.path.join(output_folder_path, f'{language}.strings')

        with open(output_file_path, 'w', encoding='utf-8') as output_file:
            # 處理每個工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                key_column = find_target_column(sheet, app_key_identifier)
                if key_column == -1:
                    print(f"\033[91mError: 'App key' column not found in sheet: {sheet_name}. Please check the column name.\033[0m")
                    os.remove(output_file_path)
                    sys.exit(1)

                target_column = find_target_column(sheet, language)
                if target_column == -1:
                    break

                # 處理工作表
                process_sheet(sheet, output_file, key_column, target_column)

        # 讀取輸出檔案
        with open(output_file_path, 'r', encoding='utf-8') as file:
            file_content = file.read()

        # 移除空白行和特定句子
        cleaned_content = remove_empty_row(file_content)
        cleaned_content = remove_empty_source(cleaned_content)

        # 只有在 cleaned_content 不是空白時才寫入檔案
        if cleaned_content.strip():  # 如果去除兩側空白後仍有字元剩餘
            # 將整理後的內容寫回檔案
            with open(output_file_path, 'w', encoding='utf-8') as file:
                file.write(cleaned_content)
        else:
            # 如果 cleaned_content 是空白，則刪除檔案
            os.remove(output_file_path)


if __name__ == "__main__":
    main()
